import os
import random

from openpyxl import load_workbook
import pymysql

from SpareFoodShare.settings import DATABASES
import datetime as dt
from django.conf import settings 
from django.contrib import auth
from django.contrib.auth.decorators import login_required
from django.core.mail import send_mail, BadHeaderError
from django.http import HttpResponseRedirect
from django.http.response import JsonResponse 
from django.http.response import JsonResponse, HttpResponse
from django.shortcuts import render, redirect
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt 
from django.views.generic.base import TemplateView
from payments.models import Food, Profile, User
from payments.userforms import LoginForm, RegistrationForm, ContactForm
import stripe
import templates.Dashboard as Dashboard


# Create your views here.
class HomePageView(TemplateView):
    template_name = 'index.html'

class SuccessView(TemplateView):
    template_name = 'success.html'

class CancelledView(TemplateView):
    template_name = 'cancelled.html'

class IndexView(TemplateView):
    template_name = 'index.html'

class ShopView(TemplateView):
    template_name = 'shoppage.html'

# class ProductView(TemplateView):
#    template_name = 'product.html'

class CollectionView(TemplateView):
    template_name = 'collection-beta.html'

class SearchView(TemplateView):
    template_name = 'search.html'

class UploadView(TemplateView):
    template_name = 'upload.html'

class PaymentView(TemplateView):
    template_name = 'payment.html'


@csrf_exempt
def stripe_config(request):
    if request.method == 'GET':
        stripe_config = {'publicKey': settings.STRIPE_PUBLISHABLE_KEY}
        return JsonResponse(stripe_config, safe=False)

@csrf_exempt
def create_checkout_session(request, pk):
    food = Food.objects.filter(foodid=pk)[0]
    print(food)
    print(food.foodname)
    print(food.quantity)
    print(food.price)
    if request.method == 'GET':
        domain_url = 'http://192.168.2.188:8000/'
        stripe.api_key = settings.STRIPE_SECRET_KEY
        try:
            # Create new Checkout Session for the order
            # Other optional params include:
            # [billing_address_collection] - to display billing address details on the page
            # [customer] - if you have an existing Stripe Customer ID
            # [payment_intent_data] - capture the payment later
            # [customer_email] - prefill the email input in the form
            # For full details see https://stripe.com/docs/api/checkout/sessions/create

            # ?session_id={CHECKOUT_SESSION_ID} means the redirect will have the session ID set as a query param
            checkout_session = stripe.checkout.Session.create(
                success_url=domain_url + 'success/' + pk,
                cancel_url=domain_url + 'cancelled/',
                payment_method_types=['card'],
                mode='payment',
                line_items=[
                    {
                        'name': food.foodname,
                        'quantity': food.quantity,
                        'currency': 'gbp',
                        'amount': int(food.price * 100),
                    }
                ]
            )
            return JsonResponse({'sessionId': checkout_session['id']})
        except Exception as e:
            return JsonResponse({'error': str(e)})

@csrf_exempt
def stripe_webhook(request):
    stripe.api_key = settings.STRIPE_SECRET_KEY
    endpoint_secret = settings.STRIPE_ENDPOINT_SECRET
    payload = request.body
    sig_header = request.META['HTTP_STRIPE_SIGNATURE']
    event = None

    try:
        event = stripe.Webhook.construct_event(
            payload, sig_header, endpoint_secret
        )
    except ValueError as e:
        # Invalid payload
        return HttpResponse(status=400)
    except stripe.error.SignatureVerificationError as e:
        # Invalid signature
        return HttpResponse(status=400)

    # Handle the checkout.session.completed event
    if event['type'] == 'checkout.session.completed':
        print("Payment was successful.")
        # TODO: run some custom code here

    return HttpResponse(status=200)

def success(request, pk):
    food = Food.objects.filter(foodid=pk)
    conn = pymysql.connect(host=DATABASES.get("default").get("HOST"), port=DATABASES.get("default").get("PORT"), user=DATABASES.get("default").get("USER"), passwd=DATABASES.get("default").get("PASSWORD"), db=DATABASES.get("default").get("NAME"),
                            charset='utf8')
    cursor = conn.cursor(cursor=pymysql.cursors.DictCursor)
    # insert into order table and remove from food
    sql2 = "INSERT INTO `order` (`Food_ID`, `User_ID`, `Order_quantity`, `Order_Amount`, `Order_Type`) " \
            "VALUES (%s,%s,%s,%s,%s)"
    order_type = 1
    try:
        p = food.price
    except:
        p = 12.13
    order_amount = 1 * int(p)
    users = User.objects.all()
    try:
        email = request.user.email
    except:
        email = "1@1.com"
    specificuser = users.filter(email=email)
    # specificuser = users.filter(email=request.user.email)
    try:
        _id = specificuser[0].id
    except:
        _id = 2

    values2 = (pk, _id, 3, order_amount, order_type)
    cursor.execute(sql2, values2)

    sql3 = "SET FOREIGN_KEY_CHECKS=0;"
    sql4 = "DELETE FROM `food` WHERE `foodID` = %s"
    sql5 = "SET FOREIGN_KEY_CHECKS=1;"

    values3 = (pk)
    cursor.execute(sql3)
    cursor.execute(sql4, values3)
    cursor.execute(sql5)
    
    conn.commit()
    conn.close()
    cursor.close()

    send_mail("food collection confirmation", "confirmation for " + "?????????", settings.DEFAULT_FROM_EMAIL, [email])

    return render(request, "success.html")


def index(request):
    foods = Food.objects.filter(expireddate__gte=timezone.now())
    print(foods)
    cat = list(set([f['categories'] for f in foods.values()]))
    # print(request.user)
    # print(request.user.is_authenticated)
    # print(cat)
    # print(foods)
    context = {
        'cat' : random.sample(cat, k=3),
        'cfoods1' : random.choices(foods, k=3),
        'cfoods2': random.choices(foods, k=3),
        'foods3' : random.choices(foods, k=4),
        'food1': random.choices(foods, k=1),
        'foods4': random.choices(foods, k=2),
        'foods5': random.choices(foods, k=2)
    }
    return render(request, "index.html", context)

def dashbroad(request):
    users = User.objects.all()
    superuser = request.user.is_superuser
    if superuser == 1:
        return render(request, "DashBroad.html")
    else:
        return HttpResponse("You Are Not A Admin! Please Return To Home Page!")

        
def shoppage(request):
    foods = Food.objects.filter(expireddate__gte=timezone.now())
    cat = list(set([f['categories'] for f in foods.values()]))
    # print(foods)
    context = {
        'cat' : random.sample(cat, k=4),
        'cfoods1' : random.choices(foods, k=3),
        'cfoods2': random.choices(foods, k=3),
        'foods3' : random.choices(foods, k=3),
        'food1': random.choices(foods, k=1),
        'foods4': random.choices(foods, k=2),
        'foods5': random.choices(foods, k=2)
    }
    return render(request, "shoppage.html", context)

def dashboard(request):
    users = User.objects.all()
    superuser = request.user.is_superuser
    if superuser == 1:
        Dashboard.CreateDashboard()
        return render(request, "DashBoard.html")
    else:
        Dashboard.CreateDashboard()
        return HttpResponse("You Are Not A Admin! Please Return To Home Page!")

def product(request, pk):
    food = Food.objects.filter(foodid=pk).values()
    fooded = Food.objects.filter(foodid=pk)
    foods = Food.objects.filter(categories__contains=food[0]['categories'], expireddate__gte=timezone.now())
    # print(fooded)
    context = {'fooded': fooded,
               'foods': foods[:8]
               }
    return render(request, "product.html", context)

# @login_required
# def upload(request):
#    print(request.user)
#    if not request.user.is_authenticated:
#        return redirect('login')
#    return render(request, "upload.html")
def search(request):
    foods = Food.objects.filter(foodname__icontains=request.POST['name'], expireddate__gte=timezone.now())
    
    context = {
        'foods': foods[:27]
    }
    return render(request, "search.html", context)

def categories(request, pk):
    foods = Food.objects.filter(categories__icontains=pk, expireddate__gte=timezone.now())
    context = {
        'foods': foods[:27]
    }
    return render(request, "categories.html", context)

def collection(request, pk):
    if not request.user.is_authenticated:
        return redirect('login')
    food = Food.objects.filter(foodid=pk).values()
    fooded = Food.objects.filter(foodid=pk)
    foods = Food.objects.filter(categories__contains=food[0]['categories'])
    # print(fooded)
    context = {'fooded': fooded,
               'foods': foods[:8]
               }
    return render(request, "collection-beta.html", context)

def payment(request, pk):
    food = Food.objects.filter(foodid=pk).values()
    fooded = Food.objects.filter(foodid=pk)
    foods = Food.objects.filter(categories__contains=food[0]['categories'])
    # print(fooded)
    context = {'fooded': fooded,
               'foods': foods[:8]
               }
    return render(request, "payment.html", context)

def logout(request):
    auth.logout(request)
    return redirect('home')

def user_login(request):
    # print(request.POST)
    if request.method == "POST":
        form = LoginForm(request.POST)
        if form.is_valid():
            username = form.cleaned_data['username']
            password = form.cleaned_data['password']


            user = auth.authenticate(username=username, password=password)
            if user is not None and user.is_active:
                auth.login(request, user)
                print(user)
                return HttpResponseRedirect(reverse("home"))
            else:
                return render(request, 'login.html',
                              {'form': form, 'message': 'The username or password is incorrect'})
        else:

            return render(request, 'login.html',
                          {'form': form, 'message': 'The username or password is incorrect'})
    else:
        form = LoginForm()

    return render(request, 'login.html', {"form": form})


def user_register(request):
    print(request.POST)
    if request.method == 'POST':
        form = RegistrationForm(request.POST)
        if form.is_valid():
            username = form.cleaned_data['username']
            email = form.cleaned_data['email']
            nickname = form.cleaned_data['nickname']
            firstname = form.cleaned_data['firstname']
            lastname = form.cleaned_data['lastname']
            password = form.cleaned_data['password']
            tel_number = form.cleaned_data['tel_number']
            gender = form.cleaned_data['gender']
            type = form.cleaned_data['type']

            user = Profile.objects.create_user(username=username,
                                               password=password,
                                               email=email,
                                               tel_number=tel_number,
                                               gender=gender,
                                               type=type,
                                               first_name=firstname,
                                               last_name=lastname,
                                               nickname=nickname
                                               )
            print(user)

            conn = pymysql.connect(host=DATABASES.get("default").get("HOST"), port=DATABASES.get("default").get("PORT"), user=DATABASES.get("default").get("USER"), passwd=DATABASES.get("default").get("PASSWORD"), db=DATABASES.get("default").get("NAME"),
                                   charset='utf8')
            cursor = conn.cursor(cursor=pymysql.cursors.DictCursor)

            sql = "INSERT INTO `user` (`FirstName`, `LastName`, `Email`, `Tel_Number`, `Gender`, `Type`) " \
          "VALUES (%s,%s,%s,%s,%s, %s)"
            values = (firstname, lastname, email, int(tel_number), gender, type)
            print(tel_number)
            # get food_id for order
            sql1 = "SELECT @@identity"
            cursor.execute(sql, values)
            cursor.execute(sql1)
            foodID = cursor.fetchall()

            conn.commit()
            conn.close()
            cursor.close()        

            return HttpResponseRedirect(reverse("login"))
    else:
        form = RegistrationForm()
    return render(request, 'register.html', {'form': form})


def time():
    year = dt.datetime.now().year
    month = dt.datetime.now().month
    day = dt.datetime.now().day
    dt2 = dt.date(year, month, day)
    new_date = dt.datetime.strptime(str(dt2), '%Y-%m-%d')
    return new_date

def upload(request):
    if not request.user.is_authenticated:
        return redirect('login')
    if request.method == "GET":
        return render(request, 'upload.html')
    else:
        redate = time()
        users = User.objects.all()
        try:
            email = request.user.email
        except:
            email = "1@1.com"
        specificuser = users.filter(email=email)
        # specificuser = users.filter(email=request.user.email)
        # specificuser = [{"id":1}]
        try:
            id = specificuser[0].id
        except:
            id = 1
            pass 
        Name = request.POST.get("Name")
        Zip_Code = request.POST.get("Zip Code")
        Category = request.POST.get("Category")
        quantity = request.POST.get("quantity")
        Location = request.POST.get("Location")
        exdate = request.POST.get("exdate")
        price = request.POST.get("price")
        Description = request.POST.get("Description")
        picture = request.FILES.get("picture")


        if len(Name) == 0 or len(Zip_Code) == 0 or len(Category) == 0 or len(quantity) == 0 or len(Location) == 0 or len(exdate) == 0 or len(price) == 0 or len(Description) == 0 or picture == None:
            return render(request, "upload.html", {"error_msg": "Please fill in the blank completely"})


        else:
            f = open("static/" + str(id) + "_" + Name + ".jpg", mode='wb+')
            for chunk in picture.chunks():
                f.write(chunk)
            f.close()
            # absolute_path = os.path.abspath(picture.name)
            absolute_path = "/static/" + str(id) + "_" + Name + ".jpg"

            conn = pymysql.connect(host=DATABASES.get("default").get("HOST"), port=DATABASES.get("default").get("PORT"), user=DATABASES.get("default").get("USER"), passwd=DATABASES.get("default").get("PASSWORD"), db=DATABASES.get("default").get("NAME"),
                                   charset='utf8')
            cursor = conn.cursor(cursor=pymysql.cursors.DictCursor)

            sql = "INSERT INTO `food` (`userID`, `foodname`, `releaseDate`, `expiredDate`, `categories`, `price`, `description`, `quantity`, `postcode`, `location`, `picture`) " \
          "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"
            values = (id, Name, redate, exdate, Category, price, Description, quantity, Zip_Code, Location, absolute_path)

            # get food_id for order
            sql1 = "SELECT @@identity"
            cursor.execute(sql, values)
            cursor.execute(sql1)
            foodID = cursor.fetchall()


            # insert into order table
            # sql2 = "INSERT INTO `order` (`Food_ID`, `User_ID`, `Order_quantity`, `Order_Amount`, `Order_Type`) " \
            #       "VALUES (%s,%s,%s,%s,%s)"
            # order_type = 1
            # order_amount=int(quantity)*int(price)

            # values2 = (foodID[0]['@@identity'], id, quantity, order_amount, order_type)
            # cursor.execute(sql2,values2)


            conn.commit()
            conn.close()
            cursor.close()

            return redirect('success')



def upload_CSV(request):


    # Now Time!
    redate = time()

    conn = pymysql.connect(host=DATABASES.get("default").get("HOST"), port=DATABASES.get("default").get("PORT"), user=DATABASES.get("default").get("USER"), passwd=DATABASES.get("default").get("PASSWORD"), db=DATABASES.get("default").get("NAME"),
                           charset='utf8')
    cursor = conn.cursor(cursor=pymysql.cursors.DictCursor)

    file_obj = request.FILES.get("csv_ex")
    try:
        wb = load_workbook(file_obj)
    except:
        return render(request, "upload.html", {"error_msg": "Please upload the csv file"})
    sheet = wb.worksheets[0]

    users = User.objects.all()
    specificuser = users.filter(email=request.user.email)
    try:
        id = specificuser[0].id
    except:
        id = 2
        pass
    for row in sheet.iter_rows(min_row=2):



        sql = "INSERT INTO `food` (`userID`, `foodname`, `releaseDate`, `expiredDate`, `categories`, `price`, `description`, `quantity`, `postcode`, `location`, `picture`) " \
              "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"

        values = (id, row[0].value, redate, row[1].value, row[2].value, row[3].value, row[4].value, row[5].value, row[6].value, row[7].value, row[8].value)


        cursor.execute(sql, values)

        sql1 = "SELECT @@identity"
        # cursor.execute(sql, values)
        cursor.execute(sql1)
        foodID = cursor.fetchall()
        conn.commit()





    conn.close()
    cursor.close()


    return redirect('success')

def contact(request, pk):
    food = Food.objects.filter(foodid=pk)[0]
    users = User.objects.all()
    print(food.userid)
    specificuser = food.userid
    if request.method == 'GET':
        form = ContactForm()
    else:
        form = ContactForm(request.POST)
        if form.is_valid():
            subject = form.cleaned_data['subject']
            try:
                from_email = request.user.email
            except:
                from_email = "1@q.com"
            message = form.cleaned_data['message']
            try:
                send_mail(subject, message, from_email, [specificuser.email])
            except BadHeaderError:
                return HttpResponse('Invalid header found.')
            return redirect('success')
    return render(request, "contact.html", {'form': form})
